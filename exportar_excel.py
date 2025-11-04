#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Módulo de Exportación a Excel
Genera reportes detallados en formato Excel
"""

import sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference

DATABASE = 'SISTEMA_CONTABLE/DATOS/contabilidad.db'

def formatear_moneda(valor):
    """Formatea un valor como moneda colombiana"""
    return f"${valor:,.0f}".replace(",", ".")

def crear_reporte_mensual(mes, anio):
    """
    Crea un reporte Excel completo del mes especificado
    """
    print(f"\n📊 Generando reporte Excel para {mes}/{anio}...")
    
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    
    # Crear workbook
    wb = Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # HOJA 1: Resumen General
    crear_hoja_resumen(wb, conn, mes, anio)
    
    # HOJA 2: Presupuestos Detallados
    crear_hoja_presupuestos(wb, conn, mes, anio)
    
    # HOJA 3: Gastos Detallados
    crear_hoja_gastos(wb, conn, mes, anio)
    
    # HOJA 4: Análisis por Categoría
    crear_hoja_analisis(wb, conn, mes, anio)
    
    conn.close()
    
    # Guardar archivo
    nombre_archivo = f"Reporte_{anio}_{mes:02d}.xlsx"
    wb.save(nombre_archivo)
    
    print(f"✅ Reporte generado: {nombre_archivo}")
    return nombre_archivo

def crear_hoja_resumen(wb, conn, mes, anio):
    """Crea hoja de resumen general"""
    ws = wb.create_sheet("📊 Resumen General", 0)
    
    # Título
    ws['A1'] = f"REPORTE FINANCIERO - {mes}/{anio}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')
    
    # Fecha de generación
    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws.merge_cells('A2:F2')
    
    # SECCIÓN: Presupuestos
    ws['A4'] = "PRESUPUESTOS DEL MES"
    ws['A4'].font = Font(size=14, bold=True)
    ws['A4'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    ws['A4'].font = Font(size=14, bold=True, color="FFFFFF")
    ws.merge_cells('A4:F4')
    
    # Obtener datos de presupuestos
    presupuestos = conn.execute('''
        SELECT categoria, etiqueta, monto_presupuestado, monto_gastado
        FROM Presupuestos
        WHERE mes = ? AND anio = ?
        ORDER BY categoria
    ''', (mes, anio)).fetchall()
    
    row = 5
    ws[f'A{row}'] = "Categoría"
    ws[f'B{row}'] = "Etiqueta"
    ws[f'C{row}'] = "Presupuestado"
    ws[f'D{row}'] = "Gastado"
    ws[f'E{row}'] = "Diferencia"
    ws[f'F{row}'] = "% Uso"
    
    # Estilo de encabezados
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    
    row += 1
    total_presupuestado = 0
    total_gastado = 0
    
    for p in presupuestos:
        ws[f'A{row}'] = p['categoria']
        ws[f'B{row}'] = p['etiqueta']
        ws[f'C{row}'] = p['monto_presupuestado']
        ws[f'D{row}'] = p['monto_gastado']
        ws[f'E{row}'] = p['monto_presupuestado'] - p['monto_gastado']
        
        porcentaje = (p['monto_gastado'] / p['monto_presupuestado'] * 100) if p['monto_presupuestado'] > 0 else 0
        ws[f'F{row}'] = f"{porcentaje:.1f}%"
        
        # Formato de moneda
        ws[f'C{row}'].number_format = '#,##0'
        ws[f'D{row}'].number_format = '#,##0'
        ws[f'E{row}'].number_format = '#,##0'
        
        total_presupuestado += p['monto_presupuestado']
        total_gastado += p['monto_gastado']
        
        row += 1
    
    # Totales
    ws[f'A{row}'] = "TOTAL"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'C{row}'] = total_presupuestado
    ws[f'D{row}'] = total_gastado
    ws[f'E{row}'] = total_presupuestado - total_gastado
    porcentaje_total = (total_gastado / total_presupuestado * 100) if total_presupuestado > 0 else 0
    ws[f'F{row}'] = f"{porcentaje_total:.1f}%"
    
    ws[f'C{row}'].number_format = '#,##0'
    ws[f'D{row}'].number_format = '#,##0'
    ws[f'E{row}'].number_format = '#,##0'
    
    # SECCIÓN: Gastos
    row += 3
    ws[f'A{row}'] = "GASTOS DEL MES"
    ws[f'A{row}'].font = Font(size=14, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
    ws.merge_cells(f'A{row}:F{row}')
    
    gastos = conn.execute('''
        SELECT COUNT(*) as total, 
               SUM(CASE WHEN estado = 'Pagado' THEN 1 ELSE 0 END) as pagados,
               SUM(CASE WHEN estado = 'Pendiente' THEN 1 ELSE 0 END) as pendientes,
               SUM(monto) as monto_total,
               SUM(CASE WHEN estado = 'Pagado' THEN monto ELSE 0 END) as monto_pagado
        FROM Egresos
        WHERE strftime('%m', fecha_vencimiento) = ? AND strftime('%Y', fecha_vencimiento) = ?
    ''', (f"{mes:02d}", str(anio))).fetchone()
    
    row += 1
    ws[f'A{row}'] = f"Total de gastos: {gastos['total']}"
    row += 1
    ws[f'A{row}'] = f"Pagados: {gastos['pagados']}"
    ws[f'C{row}'] = gastos['monto_pagado']
    ws[f'C{row}'].number_format = '#,##0'
    row += 1
    ws[f'A{row}'] = f"Pendientes: {gastos['pendientes']}"
    ws[f'C{row}'] = gastos['monto_total'] - gastos['monto_pagado']
    ws[f'C{row}'].number_format = '#,##0'
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12

def crear_hoja_presupuestos(wb, conn, mes, anio):
    """Crea hoja de presupuestos detallados"""
    ws = wb.create_sheet("💰 Presupuestos")
    
    ws['A1'] = f"PRESUPUESTOS DETALLADOS - {mes}/{anio}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:G1')
    
    # Encabezados
    headers = ['Categoría', 'Etiqueta', 'Presupuestado', 'Gastado', 'Disponible', '% Uso', 'Estado']
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    
    # Datos
    presupuestos = conn.execute('''
        SELECT * FROM Presupuestos
        WHERE mes = ? AND anio = ?
        ORDER BY categoria
    ''', (mes, anio)).fetchall()
    
    row = 4
    for p in presupuestos:
        ws.cell(row=row, column=1, value=p['categoria'])
        ws.cell(row=row, column=2, value=p['etiqueta'])
        ws.cell(row=row, column=3, value=p['monto_presupuestado']).number_format = '#,##0'
        ws.cell(row=row, column=4, value=p['monto_gastado']).number_format = '#,##0'
        ws.cell(row=row, column=5, value=p['monto_presupuestado'] - p['monto_gastado']).number_format = '#,##0'
        
        porcentaje = (p['monto_gastado'] / p['monto_presupuestado'] * 100) if p['monto_presupuestado'] > 0 else 0
        ws.cell(row=row, column=6, value=f"{porcentaje:.1f}%")
        
        if porcentaje > 100:
            estado = "❌ Excedido"
        elif porcentaje > 90:
            estado = "⚠️ Cerca"
        else:
            estado = "✅ OK"
        ws.cell(row=row, column=7, value=estado)
        
        row += 1
    
    # Ajustar columnas
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 15

def crear_hoja_gastos(wb, conn, mes, anio):
    """Crea hoja de gastos detallados"""
    ws = wb.create_sheet("📋 Gastos")
    
    ws['A1'] = f"GASTOS DETALLADOS - {mes}/{anio}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:H1')
    
    # Encabezados
    headers = ['Descripción', 'Categoría', 'Etiqueta', 'Monto', 'Vencimiento', 'Estado', 'Fecha Pago', 'Usuario']
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    
    # Datos
    gastos = conn.execute('''
        SELECT * FROM Egresos
        WHERE strftime('%m', fecha_vencimiento) = ? AND strftime('%Y', fecha_vencimiento) = ?
        ORDER BY fecha_vencimiento DESC
    ''', (f"{mes:02d}", str(anio))).fetchall()
    
    row = 4
    for g in gastos:
        ws.cell(row=row, column=1, value=g['descripcion'])
        ws.cell(row=row, column=2, value=g['categoria'])
        ws.cell(row=row, column=3, value=g['etiqueta'])
        ws.cell(row=row, column=4, value=g['monto']).number_format = '#,##0'
        ws.cell(row=row, column=5, value=g['fecha_vencimiento'])
        ws.cell(row=row, column=6, value=g['estado'])
        ws.cell(row=row, column=7, value=g['fecha_pago'] if g['fecha_pago'] else '-')
        ws.cell(row=row, column=8, value=g['usuario_que_pago'] if g['usuario_que_pago'] else '-')
        
        row += 1
    
    # Ajustar columnas
    ws.column_dimensions['A'].width = 30
    for col in range(2, 9):
        ws.column_dimensions[chr(64 + col)].width = 15

def crear_hoja_analisis(wb, conn, mes, anio):
    """Crea hoja de análisis por categoría"""
    ws = wb.create_sheet("📈 Análisis")
    
    ws['A1'] = f"ANÁLISIS POR CATEGORÍA - {mes}/{anio}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:E1')
    
    # Datos por categoría
    analisis = conn.execute('''
        SELECT 
            categoria,
            COUNT(*) as cantidad_gastos,
            SUM(monto) as total_gastos,
            AVG(monto) as promedio_gasto
        FROM Egresos
        WHERE strftime('%m', fecha_vencimiento) = ? AND strftime('%Y', fecha_vencimiento) = ?
        GROUP BY categoria
        ORDER BY total_gastos DESC
    ''', (f"{mes:02d}", str(anio))).fetchall()
    
    # Encabezados
    headers = ['Categoría', 'Cantidad', 'Total', 'Promedio', '% del Total']
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    
    total_general = sum([a['total_gastos'] for a in analisis])
    
    row = 4
    for a in analisis:
        ws.cell(row=row, column=1, value=a['categoria'])
        ws.cell(row=row, column=2, value=a['cantidad_gastos'])
        ws.cell(row=row, column=3, value=a['total_gastos']).number_format = '#,##0'
        ws.cell(row=row, column=4, value=a['promedio_gasto']).number_format = '#,##0'
        
        porcentaje = (a['total_gastos'] / total_general * 100) if total_general > 0 else 0
        ws.cell(row=row, column=5, value=f"{porcentaje:.1f}%")
        
        row += 1
    
    # Ajustar columnas
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 15

if __name__ == "__main__":
    # Generar reporte del mes actual
    ahora = datetime.now()
    crear_reporte_mensual(ahora.month, ahora.year)